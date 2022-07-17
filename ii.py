#!/usr/bin/python3
# -*- coding: utf-8 -*-
#
__author__ = 'Гаврилов Василий 91840'

from openpyxl import load_workbook
import openpyxl as opx
import pymysql

print("Откреывается dataset")
wb = load_workbook('./datas/dataset.xlsx') #,read_only=True
sheet_array = wb.sheetnames
#print(sheet_array[0])
#sheet = wb[sheet_array[0]]
ws = wb.active
# print(ws[0][1].value)

rows = ws.max_row
cols = ws.max_column

print("Строк: ", rows)
print("Столбцов: ", cols)

print("start ii")
i=0
#tnvdId = dict({})
error = 0
ws.insert_cols(11,2)

ws[1][10].value = "Ошибка! Иностранная организация должна иметь ТН ВЭД"
ws[1][10].font = opx.styles.Font(color='ff0000')
def secondFast(t,idDTnvd,idDTehr):
    try:
        codes = idDTnvd.split(';')
    except Exception:
        codes = idDTnvd
    if(isinstance(codes, int)==True):
        codes = str(codes).strip()
        try:
            if len(t[codes])>0:
                t[codes][idDTehr] += 1
        except KeyError:
            try:
                if len(t[codes])>0:
                    t[codes][idDTehr] = 1
            except KeyError:
                t[codes] = {}
                t[codes][idDTehr] = 1
    if(isinstance(codes, list)==True):
        for x in codes:
            if codes.count(x) > 1:
                codes.remove(x)
        for code in codes:
            code = str(code).strip()
            try:
                if len(t[code])>0:
                    t[code][idDTehr] += 1
            except KeyError:
                try:
                    if len(t[code])>0:
                        t[code][idDTehr] = 1
                except KeyError:
                    t[code] = {}
                    t[code][idDTehr] = 1
    return t

TnvdTehreg = {}
TnvdGroup = {}
TnvdNameprod = {}
TehregTnvd = {}
for row in ws:
    i+=1
    if i==1:
        continue
    idDTnvd = row[1].value
    idDTehr = row[2].value
    idDGroup = row[3].value
    idDName = row[4].value
    idDCity = row[9].value

    if(idDCity!="РОССИЯ")and(idDTnvd is None):
        # print(type(idDTnvd),idDTnvd)
        # print(idDCity)
        row[10].value = 1
        row[10].font = opx.styles.Font(color='ff0000')

        error+=1

    try:
        idDTehrs = idDTehr.split(';')
        for idDTehr in idDTehrs:
            idDTehr = str(idDTehr).strip()
            TnvdTehreg = secondFast(TnvdTehreg,idDTnvd,idDTehr)
    except Exception:
        idDTehr = str(idDTehr).strip()
        TnvdTehreg = secondFast(TnvdTehreg,idDTnvd,idDTehr)
    #
    try:
        idDGroups = idDGroup.split(';')
        for idDGroup in idDGroups:
            idDGroup = str(idDGroup).strip()
            TnvdGroup = secondFast(TnvdGroup,idDTnvd,idDGroup)
    except Exception:
        idDGroup = str(idDGroup).strip()
        TnvdGroup = secondFast(TnvdGroup,idDTnvd,idDGroup)

    try:
        idDNames = idDName.split(';')
        for idDName in idDNames:
            idDName = str(idDName).strip()
            TnvdNameprod = secondFast(TnvdNameprod,idDTnvd,idDName)
    except Exception:
        idDName = str(idDName).strip()
        TnvdNameprod = secondFast(TnvdNameprod,idDTnvd,idDName)



    try:
        idDTnvds = idDTnvd.split(';')
        for idDTnvd in idDTnvds:
            idDTehr = str(idDTehr).strip()
            idDTnvd = str(idDTnvd).strip()
            TehregTnvd = secondFast(TehregTnvd,idDTehr,idDTnvd)
    except Exception:
        idDTehr = str(idDTehr).strip()
        idDTnvd = str(idDTnvd).strip()
        TehregTnvd = secondFast(TehregTnvd,idDTehr,idDTnvd)


    if i%3000==1:
        print(i)
    # productNumber = ws[row][0].value
    # codestnvd = ws[row][1].value
    # technicalRegulations = ws[row][2].value
    # productGroup = ws[row][3].value
    # generalProductName = ws[row][4].value
    # il = ws[row][5].value
    # applicant = ws[row][6].value
    # applicantAddress = ws[row][7].value
    # manufacturer = ws[row][8].value
    # country = ws[row][9].value
    # manufacturerAddress = ws[row][10].value
print("end")
print("Ошибок по стране без ТНВЭД",error)
print("Кодов ТНВД",len(TnvdTehreg))
print("9503004900 TnvdTehreg: ",TnvdTehreg["9503004900"])
print("Групп товаров",len(TnvdGroup))
print("9503004900 TnvdGroup: ",TnvdGroup["9503004900"])
# print("9503004900 TnvdNameprod: ",TnvdNameprod["9503004900"])
print("9503004900 TehregTnvd: ",TehregTnvd["ТР ТС 008/2011 О безопасности игрушек"])

print("950300100 TnvdNameprod: ",TnvdNameprod["950300100"])

# 'ТР ЕАЭС 043/2017 О требованиях к средствам обеспечения пожарной безопасности и пожаротушения'
# ws.template = False , as_template=False
wb.save('./datas/document.xlsx')
print("Сохренен!")
