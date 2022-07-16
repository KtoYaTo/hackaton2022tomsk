#!/usr/bin/python3
# -*- coding: utf-8 -*-
#
__author__ = 'Гаврилов Василий 91840'

from openpyxl import load_workbook
import pymysql

print("Откреывается dataset")
wb = load_workbook('./datas/dataset.xlsx',read_only=True)
sheet_array = wb.sheetnames
#print(sheet_array[0])
#sheet = wb[sheet_array[0]]
ws = wb.active
# print(ws[0][1].value)

rows = ws.max_row
cols = ws.max_column

print("Строк: ", rows)
print("Столбцов: ", cols)

conn = pymysql.connect(host='37.140.192.237', user='u0863665_hack', password='cY7qF7rW7p', database='u0863665_hack', charset='utf8mb4', cursorclass=pymysql.cursors.DictCursor)



# Номер продукции
# Коды ТН ВЭД ЕАЭС
# Технические регламенты
# Группа продукции
# Общее наименование продукции
# ИЛ
# Заявитель
# Адрес Заявителя
# Изготовитель
# Страна
# Адрес изготовителя
try:
    with conn:
        cur = conn.cursor()
        #cur.execute("SELECT * FROM cities")
        #rows = cur.fetchall()
        cur.execute("TRUNCATE dataset")
        cur.execute("TRUNCATE tndv_codes")
        conn.commit()
        i=0
        for row in ws:
            i+=1
            if i==1:
                continue
            if i%100==1:
                print(i)
                conn.commit()
            # for col in range(0,cols):
            # print(ws[row][col].value)
            productNumber = row[0].value
            productNumber = str(int(productNumber.replace("Продукция", "").strip()))
            codestnvd = row[1].value
            technicalRegulations = row[2].value
            productGroup = row[3].value
            generalProductName = row[4].value
            il = row[5].value
            applicant = row[6].value
            applicantAddress = row[7].value
            manufacturer = row[8].value
            country = row[9].value
            manufacturerAddress = row[10].value
            sqlQuery = "INSERT INTO `dataset` SET `productNumber`="+productNumber+", `techReg`=%s, `productGroup`=%s, `generalProductName`=%s, `il`=%s, `applicant`=%s, `applicantAddress`=%s, `manufacturer`=%s, `country`=%s, `manufacturerAddress`=%s;"
            data = (technicalRegulations,productGroup,generalProductName,il,applicant,applicantAddress,manufacturer,country,manufacturerAddress)
            cur.execute(sqlQuery,data)
            # conn.commit()

            dataset_id = cur.lastrowid
            # conn.insert_id()
            # print(dataset_id)
            if(codestnvd!=""):
                try:
                    codes = codestnvd.split(';')
                except Exception:
                    codes = codestnvd

                if(isinstance(codes, int)==True):
                    sqlQuery = "INSERT INTO tndv_codes SET `code`="+str(codes)+", `dataset_id`="+str(dataset_id)+";"
                    cur.execute(sqlQuery)
                if(isinstance(codes, list)==True):
                    for code in codes:
                        sqlQuery = "INSERT INTO tndv_codes SET `code`="+str(code).strip()+", `dataset_id`="+str(dataset_id)+";"
                        cur.execute(sqlQuery)
        print("отправка")
        conn.commit()


except Exception as e:
    print("Ошбика:{}".format(e))
# finally:
#     conn.close()